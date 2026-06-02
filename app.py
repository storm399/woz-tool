"""WOZ-waardeloket webtool.

Lokaal Flask-webserver dat:
- PDOK Locatieserver gebruikt om adres -> nummeraanduiding te resolven
- WOZ-waardeloket API (api.kadaster.nl) aanroept om WOZ-waarden op te vragen
- Enkel adres in browser laat invoeren OF batch Excel upload + download

Start: python3 app.py  ->  open http://127.0.0.1:5005
"""

import io
import os
import time
import re
from typing import Optional

import requests
from flask import Flask, render_template, request, jsonify, send_file, abort

import realworks_service
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule


def _load_dotenv(path: str = ".env") -> None:
    """Mini-loader voor .env (alleen lokaal; op Render zet je env vars in dashboard)."""
    if not os.path.exists(path):
        return
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


_load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

app = Flask(__name__)

PDOK_FREE = "https://api.pdok.nl/bzk/locatieserver/search/v3_1/free"
PDOK_SUGGEST = "https://api.pdok.nl/bzk/locatieserver/search/v3_1/suggest"
WOZ_API = "https://api.kadaster.nl/lvwoz/wozwaardeloket-api/v1/wozwaarde/nummeraanduiding/{}"
WOZ_HEADERS = {
    "Accept": "application/json",
    "Origin": "https://www.wozwaardeloket.nl",
    "Referer": "https://www.wozwaardeloket.nl/",
    "User-Agent": "Mozilla/5.0 (WOZ-tool lokaal)",
}

EPO_API = "https://public.ep-online.nl/api/v5/PandEnergielabel/AdresseerbaarObject/{}"
EPO_API_KEY = os.environ.get("EPONLINE_API_KEY", "").strip()

BAG_WFS = "https://service.pdok.nl/lv/bag/wfs/v2_0"
CBS_WFS = "https://service.pdok.nl/cbs/wijkenbuurten/2023/wfs/v1_0"
RCE_SPARQL = "https://api.linkeddata.cultureelerfgoed.nl/datasets/rce/cho/sparql"

REQUEST_TIMEOUT = 15
session = requests.Session()


def _normaliseer(s: str) -> str:
    """Normaliseert adres-string voor match-vergelijking."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _check_adres_match(invoer: str, gevonden_straat: str, gevonden_huisnummer) -> tuple:
    """Returnt (match: bool, reden: str).

    Match = True als de gevonden straatnaam (genormaliseerd) als substring
    in de invoer voorkomt EN het gevonden huisnummer in de invoer staat.
    """
    inv = _normaliseer(invoer)
    if not inv:
        return False, "Adres leeg"

    straat_norm = _normaliseer(gevonden_straat or "")
    nr = str(gevonden_huisnummer) if gevonden_huisnummer is not None else ""

    straat_ok = bool(straat_norm) and straat_norm in inv
    if not straat_ok:
        return False, f"Straat ‘{gevonden_straat}’ niet in invoer"

    if nr:
        invoer_nums = re.findall(r"\d+", inv)
        if nr not in invoer_nums:
            return False, f"Huisnummer {nr} niet in invoer"

    return True, "OK"


# ---------------------------------------------------------------------------
# Strikte adresmatching (gebruikt door Excel-batch en Realworks-verrijking)
# ---------------------------------------------------------------------------

POSTCODE_RE = re.compile(r"(\d{4})\s*([A-Za-z]{2})")


def _normaliseer_postcode(raw) -> Optional[str]:
    """Normaliseert een postcode-invoer naar het formaat '1234AB'.

    - Spaties worden verwijderd, letters worden hoofdletters.
    - Werkt zowel op een 'kale' postcode ('1234 ab') als op een adresstring
      waar de postcode ergens in voorkomt ('Vondelstraat 70, 1054 GG Amsterdam').
    Returnt None als er geen geldige postcode in de invoer staat.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Eerst: hele string als 'kale' postcode behandelen (spaties weg, uppercase)
    compact = s.replace(" ", "").upper()
    if re.fullmatch(r"\d{4}[A-Z]{2}", compact):
        return compact
    # Anders: ergens in de string zoeken
    m = POSTCODE_RE.search(s)
    if m:
        return (m.group(1) + m.group(2)).upper()
    return None


def _split_huisnummer_componenten(raw) -> tuple:
    """Splitst een ruwe huisnummer-invoer in (nummer:int, huisletter:str|None, toevoeging:str|None).

    Voorbeelden:
        "70"        -> (70, None, None)
        "70a"       -> (70, "A", None)
        "70 A"      -> (70, "A", None)
        "70-3"      -> (70, None, "3")
        "70a-bis"   -> (70, "A", "bis")
        "70 huis"   -> (70, None, "huis")
    Returnt (None, None, None) als er geen cijferreeks te vinden is.
    """
    if raw is None:
        return None, None, None
    s = str(raw).strip()
    if not s:
        return None, None, None
    m = re.match(r"^\s*(\d+)\s*(.*)$", s)
    if not m:
        return None, None, None
    nummer = int(m.group(1))
    rest = m.group(2).strip()
    # Voorloop-streepjes en spaties wegfilteren
    rest = re.sub(r"^[\s\-]+", "", rest)
    if not rest:
        return nummer, None, None
    # Eerste teken een losse letter (gevolgd door eind, spatie of streepje)? -> huisletter
    m2 = re.match(r"^([A-Za-z])(?:[\s\-]+(.*))?$", rest)
    if m2:
        huisletter = m2.group(1).upper()
        toev = (m2.group(2) or "").strip()
        return nummer, huisletter, (toev or None)
    return nummer, None, rest


def _huisnr_uit_adresstring(adres: str, postcode: Optional[str] = None) -> tuple:
    """Probeert huisnummer + huisletter + toevoeging uit een vrij adresveld te halen.

    Strategie: knip de string af bij de postcode (indien aanwezig); zoek dan het
    láátste cijferblok in het overgebleven 'straat + nummer' deel.
    """
    if not adres:
        return None, None, None
    s = str(adres).strip()
    if not s:
        return None, None, None
    if postcode:
        pc_pos = POSTCODE_RE.search(s)
        if pc_pos:
            s = s[: pc_pos.start()].rstrip(" ,;-")
    # Laatste 'cijfers + optionele letter + optionele toevoeging' aan het einde van de string
    m = re.search(
        r"(\d+)\s*([A-Za-z])?\s*(?:[-\s]+([A-Za-z0-9]+))?\s*$",
        s,
    )
    if not m:
        return None, None, None
    nummer = int(m.group(1))
    letter = m.group(2).upper() if m.group(2) else None
    toev = m.group(3).strip() if m.group(3) else None
    # Voorkom dat een 4-cijferig blok dat eigenlijk deel van een postcode is wordt gepakt
    if nummer >= 1000 and len(str(nummer)) == 4 and not letter and not toev:
        # Mogelijk pakte hij een onbedoeld nummer (komt eigenlijk niet voor na strip,
        # maar veilig is veilig). Probeer dan een eerder cijfer.
        m2 = re.search(r"(\d{1,4})\s*([A-Za-z])?\s*(?:[-\s]+([A-Za-z0-9]+))?\s+\S+\s*$", s)
        if m2:
            try:
                nummer = int(m2.group(1))
                letter = m2.group(2).upper() if m2.group(2) else None
                toev = m2.group(3).strip() if m2.group(3) else None
            except ValueError:
                pass
    return nummer, letter, toev


def zoek_adres_strict(
    postcode: str,
    huisnummer: int,
    huisletter: Optional[str] = None,
    toevoeging: Optional[str] = None,
    straat_hint: Optional[str] = None,
) -> Optional[dict]:
    """Strikte adres-lookup op postcode + huisnummer (+huisletter, +toevoeging).

    Gebruikt PDOK Locatieserver met fq-filters zodat we exact die ene
    nummeraanduiding krijgen — geen valse matches op alleen straatnaam.

    Returnt het PDOK-doc bij een eenduidige exacte match, anders None.
    """
    pc = _normaliseer_postcode(postcode)
    if not pc or huisnummer is None:
        return None

    fq = ["type:adres", f"postcode:{pc}", f"huisnummer:{huisnummer}"]
    try:
        r = session.get(
            PDOK_FREE,
            params={"q": "*:*", "fq": fq, "rows": 50},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        docs = r.json().get("response", {}).get("docs", []) or []
    except requests.RequestException:
        return None

    if not docs:
        return None

    def _norm(v):
        if v is None:
            return None
        s = str(v).strip().upper()
        return s or None

    want_letter = _norm(huisletter)
    want_toev = _norm(toevoeging)

    def matcht(d):
        d_letter = _norm(d.get("huisletter"))
        d_toev = _norm(d.get("huisnummertoevoeging"))
        return d_letter == want_letter and d_toev == want_toev

    exact = [d for d in docs if matcht(d)]

    # Geen exacte match? Strikt: geen guess.
    if not exact:
        return None

    if len(exact) == 1:
        return exact[0]

    # Meerdere matches met dezelfde huisletter+toevoeging (zelden, maar mogelijk
    # bij split-objecten). Probeer te disambigueren via straat-hint.
    if straat_hint:
        sh = _normaliseer(straat_hint)
        gefilterd = [d for d in exact if _normaliseer(d.get("straatnaam") or "") == sh]
        if len(gefilterd) == 1:
            return gefilterd[0]

    # Nog steeds ambigu -> liever geen match dan een foutieve.
    return None


def _bereken_afgeleiden(res: dict) -> None:
    """Vult res in-place met afgeleide velden: WOZ/m², %-stijging, match-flag."""
    waarden = res.get("wozWaarden") or []
    bag = res.get("bag") or {}
    opp = bag.get("gebruiksoppervlakte")
    laatste = waarden[0]["vastgesteldeWaarde"] if waarden and waarden[0].get("vastgesteldeWaarde") else None

    res["woz_per_m2"] = int(round(laatste / opp)) if (laatste and opp) else None

    # Stijgingspercentages
    if len(waarden) >= 2 and waarden[0].get("vastgesteldeWaarde") and waarden[1].get("vastgesteldeWaarde"):
        cur = waarden[0]["vastgesteldeWaarde"]
        prev = waarden[1]["vastgesteldeWaarde"]
        res["pct_1jr"] = round((cur - prev) / prev * 100, 1)
    else:
        res["pct_1jr"] = None

    def vergelijk_x_jaar_geleden(jaren):
        if not waarden or not laatste:
            return None
        peildatum_doel = None
        from datetime import datetime
        try:
            cur_year = int(waarden[0]["peildatum"][:4])
        except (KeyError, ValueError, TypeError):
            return None
        doel_year = cur_year - jaren
        for w in waarden:
            if w.get("peildatum", "").startswith(str(doel_year)):
                doel = w.get("vastgesteldeWaarde")
                if doel:
                    return round((laatste - doel) / doel * 100, 1)
        # Anders: pak oudst beschikbaar
        oudste = waarden[-1].get("vastgesteldeWaarde") if waarden else None
        return round((laatste - oudste) / oudste * 100, 1) if oudste else None

    res["pct_5jr"] = vergelijk_x_jaar_geleden(5)

    # Vergelijking met oudst beschikbaar (vaak 2014)
    if waarden and waarden[-1].get("vastgesteldeWaarde") and laatste:
        oudste = waarden[-1]["vastgesteldeWaarde"]
        res["pct_sinds_oudst"] = round((laatste - oudste) / oudste * 100, 1)
        res["oudste_peiljaar"] = (waarden[-1].get("peildatum") or "")[:4]
    else:
        res["pct_sinds_oudst"] = None
        res["oudste_peiljaar"] = None

    # Adres-match-check: gebruik losse straat+huisnummer ipv hele weergavenaam
    bag_straat = res.get("straat") or (res.get("bag") or {}).get("straat")
    bag_nr = res.get("huisnummer")
    match, reden = _check_adres_match(res.get("adres_invoer", ""), bag_straat, bag_nr)
    res["adres_match"] = match
    res["adres_match_reden"] = reden


def zoek_adres(query: str) -> Optional[dict]:
    """Zoekt een adres via PDOK Locatieserver. Returnt eerste hit of None."""
    try:
        r = session.get(
            PDOK_FREE,
            params={"q": query, "fq": "type:adres", "rows": 1},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        docs = r.json().get("response", {}).get("docs", [])
        return docs[0] if docs else None
    except requests.RequestException:
        return None


def haal_woz(nummeraanduiding_id: str) -> Optional[dict]:
    """Haalt WOZ-waarden op via Kadaster-API. Returnt dict of None bij 404."""
    try:
        r = session.get(
            WOZ_API.format(nummeraanduiding_id),
            headers=WOZ_HEADERS,
            timeout=REQUEST_TIMEOUT,
        )
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.json()
    except requests.RequestException:
        return None


def haal_bag(adresseerbaar_object_id: str) -> Optional[dict]:
    """Haalt BAG-verblijfsobject op via PDOK WFS (geen API-key nodig).

    Returnt dict met bouwjaar, gebruiksoppervlakte (m²), gebruiksdoel,
    pandidentificatie en statussen.
    """
    if not adresseerbaar_object_id:
        return None
    fil = (
        f"<Filter><PropertyIsEqualTo>"
        f"<PropertyName>bag:identificatie</PropertyName>"
        f"<Literal>{adresseerbaar_object_id}</Literal>"
        f"</PropertyIsEqualTo></Filter>"
    )
    try:
        r = session.get(
            BAG_WFS,
            params={
                "service": "WFS",
                "version": "2.0.0",
                "request": "GetFeature",
                "typeNames": "bag:verblijfsobject",
                "outputFormat": "application/json",
                "srsName": "EPSG:4326",
                "filter": fil,
            },
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return None
        p = feats[0].get("properties", {}) or {}
        return {
            "bouwjaar": p.get("bouwjaar"),
            "gebruiksoppervlakte": p.get("oppervlakte"),
            "gebruiksdoel": p.get("gebruiksdoel"),
            "vbo_status": p.get("status"),
            "pand_id": p.get("pandidentificatie"),
            "pand_status": p.get("pandstatus"),
            "straat": p.get("openbare_ruimte"),
            "huisnummer": p.get("huisnummer"),
        }
    except (requests.RequestException, ValueError):
        return None


def haal_cbs_buurt(buurtcode: str) -> Optional[dict]:
    """Haalt CBS-kerncijfers voor een buurt op via PDOK Wijken-en-Buurten WFS.

    Verwacht buurtcode in formaat 'BU05990112'. Eenheden:
    - gemiddeldeWoningwaarde: in duizenden euro's (we vermenigvuldigen ×1000)
    - gemiddeldInkomenPerInkomensontvanger: idem ×1000
    """
    if not buurtcode:
        return None
    fil = (
        f"<Filter><PropertyIsEqualTo>"
        f"<PropertyName>buurtcode</PropertyName>"
        f"<Literal>{buurtcode}</Literal>"
        f"</PropertyIsEqualTo></Filter>"
    )
    try:
        r = session.get(
            CBS_WFS,
            params={
                "service": "WFS",
                "version": "2.0.0",
                "request": "GetFeature",
                "typeNames": "wijkenbuurten:buurten",
                "outputFormat": "application/json",
                "srsName": "EPSG:4326",
                "filter": fil,
                "count": 1,
            },
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        feats = r.json().get("features", [])
        if not feats:
            return None
        p = feats[0].get("properties", {}) or {}

        def _g(key, mult=1):
            v = p.get(key)
            if v is None or (isinstance(v, (int, float)) and v <= -99000):
                return None
            return v * mult if isinstance(v, (int, float)) else v

        return {
            "buurtnaam": _g("buurtnaam"),
            "gemeentenaam": _g("gemeentenaam"),
            "aantal_inwoners": _g("aantalInwoners"),
            "bevolkingsdichtheid": _g("bevolkingsdichtheidInwonersPerKm2"),
            "gem_woningwaarde": _g("gemiddeldeWoningwaarde", 1000),
            "gem_inkomen": _g("gemiddeldInkomenPerInkomensontvanger", 1000),
            "pct_huur": _g("percentageHuurwoningen"),
            "pct_koop": _g("percentageKoopwoningen"),
            "huishoudgrootte": _g("gemiddeldeHuishoudsgrootte"),
            "woningvoorraad": _g("woningvoorraad"),
        }
    except (requests.RequestException, ValueError):
        return None


def haal_monument(pand_id: str) -> Optional[dict]:
    """Checkt of een BAG-pand een rijksmonument is via RCE Linked Data SPARQL.

    Returnt dict met monumentnummer + adres uit RCE, of None.
    """
    if not pand_id:
        return None
    query = (
        "PREFIX ceo: <https://linkeddata.cultureelerfgoed.nl/def/ceo#>\n"
        "SELECT ?nummer ?straat ?huisnr WHERE {\n"
        "  ?monument a ceo:Rijksmonument ;\n"
        "            ceo:rijksmonumentnummer ?nummer ;\n"
        "            ceo:heeftBasisregistratieRelatie/ceo:heeftBAGRelatie ?bagrel .\n"
        f"  ?bagrel ceo:pandIdentificatie \"{pand_id}\" .\n"
        "  OPTIONAL { ?bagrel ceo:openbareRuimte ?straat ; ceo:huisnummer ?huisnr }\n"
        "} LIMIT 1"
    )
    try:
        r = session.get(
            RCE_SPARQL,
            params={"query": query, "format": "json"},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        rows = r.json()
        if not rows:
            return None
        row = rows[0]
        return {
            "monumentnummer": row.get("nummer"),
            "monument_straat": row.get("straat"),
            "monument_huisnummer": row.get("huisnr"),
        }
    except (requests.RequestException, ValueError):
        return None


def haal_energielabel(adresseerbaar_object_id: str) -> Optional[dict]:
    """Haalt het meest recente energielabel op via EP-online.

    Returnt dict met velden energieklasse, registratiedatum, geldig_tot,
    bouwjaar, gebouwtype — of None als er geen label is.
    """
    if not EPO_API_KEY:
        return None
    try:
        r = session.get(
            EPO_API.format(adresseerbaar_object_id),
            headers={"Authorization": EPO_API_KEY, "Accept": "application/json"},
            timeout=REQUEST_TIMEOUT,
        )
        if r.status_code == 404:
            return None
        r.raise_for_status()
        data = r.json()
        if not data:
            return None
        # Meest recente label (hoogste registratiedatum)
        labels = sorted(
            data,
            key=lambda x: x.get("Registratiedatum") or "",
            reverse=True,
        )
        latest = labels[0]
        return {
            "energieklasse": latest.get("Energieklasse"),
            "registratiedatum": (latest.get("Registratiedatum") or "")[:10],
            "geldig_tot": (latest.get("Geldig_tot") or "")[:10],
            "bouwjaar": latest.get("Bouwjaar"),
            "gebouwtype": latest.get("Gebouwtype"),
            "gebouwklasse": latest.get("Gebouwklasse"),
            "berekend_energieverbruik": latest.get("BerekendeEnergieverbruik"),
        }
    except requests.RequestException:
        return None


def _resultaat_van_hit(adres_invoer: str, hit: dict, methode: str = "fuzzy") -> dict:
    """Bouwt een volledig WOZ/BAG/CBS/label-resultaat op vanaf een al-opgehaalde PDOK-hit.

    Wordt zowel door de single-adres flow ('fuzzy') als de Excel/Realworks flow
    ('strict') gebruikt. Doet ZELF géén adres-lookup meer.
    """
    weergavenaam = hit.get("weergavenaam", "")
    nummeraanduiding = hit.get("nummeraanduiding_id")
    adresseerbaar_object_id = hit.get("adresseerbaarobject_id")
    if not nummeraanduiding:
        return {
            "adres_invoer": adres_invoer,
            "weergavenaam": weergavenaam,
            "status": "Geen nummeraanduiding",
            "wozWaarden": [],
            "match_methode": methode,
        }

    buurtcode = hit.get("buurtcode")
    buurtnaam = hit.get("buurtnaam")

    woz = haal_woz(nummeraanduiding)
    bag = haal_bag(adresseerbaar_object_id) if adresseerbaar_object_id else None
    energielabel = haal_energielabel(adresseerbaar_object_id) if adresseerbaar_object_id else None
    cbs = haal_cbs_buurt(buurtcode) if buurtcode else None
    monument = haal_monument(bag.get("pand_id") if bag else None)

    if not woz:
        res_geen = {
            "adres_invoer": adres_invoer,
            "weergavenaam": weergavenaam,
            "nummeraanduiding": nummeraanduiding,
            "adresseerbaarobject_id": adresseerbaar_object_id,
            "buurtnaam": buurtnaam,
            "buurtcode": buurtcode,
            "cbs": cbs,
            "monument": monument,
            "status": "Geen WOZ-waarde bekend (niet-woning of onbekend)",
            "wozWaarden": [],
            "bag": bag,
            "energielabel": energielabel,
            "match_methode": methode,
        }
        _bereken_afgeleiden(res_geen)
        return res_geen

    obj = woz.get("wozObject", {}) or {}
    waarden = woz.get("wozWaarden", []) or []

    res = {
        "adres_invoer": adres_invoer,
        "weergavenaam": weergavenaam,
        "nummeraanduiding": nummeraanduiding,
        "adresseerbaarobject_id": adresseerbaar_object_id,
        "buurtnaam": buurtnaam,
        "buurtcode": buurtcode,
        "status": "OK",
        "wozobjectnummer": obj.get("wozobjectnummer"),
        "straat": obj.get("openbareruimtenaam"),
        "huisnummer": obj.get("huisnummer"),
        "huisletter": obj.get("huisletter"),
        "huisnummertoevoeging": obj.get("huisnummertoevoeging"),
        "postcode": obj.get("postcode"),
        "woonplaats": obj.get("woonplaatsnaam"),
        "grondoppervlakte": obj.get("grondoppervlakte"),
        "wozWaarden": sorted(
            [
                {"peildatum": w.get("peildatum"), "vastgesteldeWaarde": w.get("vastgesteldeWaarde")}
                for w in waarden
            ],
            key=lambda x: x["peildatum"] or "",
            reverse=True,
        ),
        "bag": bag,
        "energielabel": energielabel,
        "cbs": cbs,
        "monument": monument,
        "match_methode": methode,
    }
    _bereken_afgeleiden(res)
    return res


def maak_resultaat(adres: str) -> dict:
    """Fuzzy adres-lookup (gebruikt door de handmatige zoekfunctie).

    Roept PDOK 'free search' met één query en pakt het eerste resultaat. Goed
    voor de single-search in de UI waar de gebruiker de zoekterm zelf typt
    en uit de autocomplete-suggesties kan kiezen.
    """
    hit = zoek_adres(adres)
    if not hit:
        return {
            "adres_invoer": adres,
            "status": "Adres niet gevonden",
            "wozWaarden": [],
            "match_methode": "fuzzy",
        }
    return _resultaat_van_hit(adres, hit, methode="fuzzy")


def maak_resultaat_strict(
    adres_invoer: str,
    postcode: Optional[str],
    huisnummer: Optional[int],
    huisletter: Optional[str] = None,
    toevoeging: Optional[str] = None,
    straat_hint: Optional[str] = None,
) -> dict:
    """Strikte adres-lookup op postcode + huisnummer (+ huisletter, toevoeging).

    Geeft alleen een resultaat terug bij een eenduidige exacte match.
    Bij ontbrekende postcode, ontbrekend huisnummer of geen exacte match in PDOK
    krijg je een resultaat met status/adres_match_reden zodat je dat in de
    uitvoer kunt tonen — liever geen koppeling dan een foutieve.
    """
    pc = _normaliseer_postcode(postcode) if postcode else None
    if postcode and not pc:
        return {
            "adres_invoer": adres_invoer,
            "status": f"Ongeldige postcode: '{postcode}'",
            "wozWaarden": [],
            "adres_match": False,
            "adres_match_reden": f"Ongeldige postcode: '{postcode}'",
            "match_methode": "strict",
            "postcode_invoer": str(postcode),
            "huisnummer_invoer": huisnummer,
        }
    if not pc:
        return {
            "adres_invoer": adres_invoer,
            "status": "Postcode ontbreekt — geen exacte match mogelijk",
            "wozWaarden": [],
            "adres_match": False,
            "adres_match_reden": "Postcode ontbreekt",
            "match_methode": "strict",
            "huisnummer_invoer": huisnummer,
        }
    if huisnummer is None:
        return {
            "adres_invoer": adres_invoer,
            "status": "Huisnummer ontbreekt — geen exacte match mogelijk",
            "wozWaarden": [],
            "adres_match": False,
            "adres_match_reden": "Huisnummer ontbreekt",
            "match_methode": "strict",
            "postcode_invoer": pc,
        }

    hit = zoek_adres_strict(pc, huisnummer, huisletter, toevoeging, straat_hint)
    if not hit:
        beschrijving = f"{pc} {huisnummer}"
        if huisletter:
            beschrijving += huisletter
        if toevoeging:
            beschrijving += f"-{toevoeging}"
        return {
            "adres_invoer": adres_invoer,
            "status": f"Geen exacte match voor {beschrijving}",
            "wozWaarden": [],
            "adres_match": False,
            "adres_match_reden": (
                f"Geen exacte match op postcode {pc} + huisnummer {huisnummer}"
                + (f" {huisletter}" if huisletter else "")
                + (f"-{toevoeging}" if toevoeging else "")
            ),
            "match_methode": "strict",
            "postcode_invoer": pc,
            "huisnummer_invoer": huisnummer,
        }

    res = _resultaat_van_hit(adres_invoer, hit, methode="strict")
    # Bij strict-match is de adres-match definitionally OK (postcode + huisnummer
    # zijn al exact gecheckt). De fuzzy substring-check in _bereken_afgeleiden
    # hier overschrijven zodat een ingevoerde 'kale' postcode+nummer niet
    # onterecht als mismatch verschijnt.
    res["adres_match"] = True
    res["adres_match_reden"] = "Exacte match op postcode + huisnummer"
    return res


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/suggest")
def api_suggest():
    """Auto-suggest endpoint voor het zoekveld."""
    q = (request.args.get("q") or "").strip()
    if len(q) < 3:
        return jsonify([])
    try:
        r = session.get(
            PDOK_SUGGEST,
            params={"q": q, "fq": "type:adres", "rows": 8},
            timeout=REQUEST_TIMEOUT,
        )
        r.raise_for_status()
        docs = r.json().get("response", {}).get("docs", [])
        return jsonify([{"id": d.get("id"), "weergavenaam": d.get("weergavenaam")} for d in docs])
    except requests.RequestException as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/woz")
def api_woz():
    """Single-address lookup endpoint."""
    adres = (request.args.get("adres") or "").strip()
    if not adres:
        return jsonify({"error": "Parameter 'adres' is verplicht"}), 400
    return jsonify(maak_resultaat(adres))


@app.route("/api/excel", methods=["POST"])
def api_excel():
    """Batch-verwerking: Excel in -> Excel out met WOZ-data.

    STRIKTE adresmatching: postcode + huisnummer (+ optioneel huisletter en
    toevoeging) is leidend. Straat/plaats worden alleen als hint gebruikt voor
    disambiguatie bij meerdere matches. Bij ontbrekende postcode, ontbrekend
    huisnummer of geen exacte match in PDOK krijgt de rij een duidelijke
    foutmelding — er wordt nooit op alleen straatnaam gegokt.
    """
    if "file" not in request.files:
        abort(400, "Geen bestand geüpload (veldnaam: 'file').")
    upload = request.files["file"]
    if not upload.filename.lower().endswith((".xlsx", ".xlsm")):
        abort(400, "Alleen .xlsx of .xlsm bestanden worden ondersteund.")

    try:
        wb_in = load_workbook(io.BytesIO(upload.read()), data_only=True)
    except Exception as e:
        abort(400, f"Kan Excel niet lezen: {e}")

    ws_in = wb_in.active
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        abort(400, "Het Excel-bestand is leeg.")

    header = [str(c) if c is not None else "" for c in rows[0]]
    header_lower = [h.lower().strip() for h in header]

    def _find(*opties):
        for i, h in enumerate(header_lower):
            if h in opties:
                return i
        return None

    adres_idx = _find(
        "adres", "address", "weergavenaam", "volledig adres", "volledig_adres",
    )
    straat_idx = _find("straat", "straatnaam", "openbareruimte", "openbare_ruimte")
    huisnummer_idx = _find("huisnummer", "nr", "nummer", "huisnr", "huis nr")
    huisletter_idx = _find("huisletter", "letter")
    toevoeging_idx = _find(
        "toevoeging", "huisnummertoevoeging", "huisnr toevoeging",
        "huis toevoeging", "huisnummer toevoeging",
    )
    postcode_idx = _find("postcode", "pc", "postal code", "zipcode", "zip")
    plaats_idx = _find("plaats", "woonplaats", "stad", "gemeente")

    def _cel(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        if isinstance(v, str) and not v.strip():
            return None
        return v

    def extract_componenten(row) -> dict:
        """Bouwt postcode/huisnummer/huisletter/toevoeging op uit een rij.

        Pakt eerst expliciete kolommen; valt anders terug op het 'adres'-veld
        en parsed daar postcode + huisnummer uit met regex.
        Returnt ook een 'adres_invoer' weergavestring voor logging.
        """
        pc_raw = _cel(row, postcode_idx)
        hn_raw = _cel(row, huisnummer_idx)
        hl_raw = _cel(row, huisletter_idx)
        tv_raw = _cel(row, toevoeging_idx)
        str_raw = _cel(row, straat_idx)
        pl_raw = _cel(row, plaats_idx)
        adres_raw = _cel(row, adres_idx)

        # Weergave-string voor logging/output
        if adres_raw:
            adres_invoer = str(adres_raw).strip()
        else:
            parts = []
            if str_raw:
                parts.append(str(str_raw).strip())
            if hn_raw is not None:
                parts.append(str(hn_raw).strip())
            if hl_raw:
                parts.append(str(hl_raw).strip())
            if tv_raw:
                parts.append(f"-{str(tv_raw).strip()}")
            if pc_raw:
                parts.append(str(pc_raw).strip())
            if pl_raw:
                parts.append(str(pl_raw).strip())
            adres_invoer = " ".join(parts).strip()

        # Postcode: kolom > parse uit adres-string
        postcode = _normaliseer_postcode(pc_raw) if pc_raw else None
        if not postcode and adres_raw:
            postcode = _normaliseer_postcode(str(adres_raw))

        # Huisnummer + huisletter + toevoeging
        huisnr = None
        huisletter = None
        toevoeging = None
        if hn_raw is not None:
            # Combineer kolommen — soms zit alles in 'huisnummer' (bv. "70a")
            samengevoegd = str(hn_raw).strip()
            if hl_raw:
                samengevoegd += f" {str(hl_raw).strip()}"
            if tv_raw:
                samengevoegd += f"-{str(tv_raw).strip()}"
            huisnr, huisletter, toevoeging = _split_huisnummer_componenten(samengevoegd)
        elif adres_raw:
            huisnr, huisletter, toevoeging = _huisnr_uit_adresstring(
                str(adres_raw), postcode
            )

        # Aparte huisletter/toevoeging kolommen winnen van wat we uit een
        # samengevoegde string haalden, als ze expliciet ingevuld zijn.
        if hl_raw and not huisletter:
            huisletter = str(hl_raw).strip().upper() or None
        if tv_raw and not toevoeging:
            toevoeging = str(tv_raw).strip() or None

        return {
            "adres_invoer": adres_invoer,
            "postcode": postcode,
            "huisnummer": huisnr,
            "huisletter": huisletter,
            "toevoeging": toevoeging,
            "straat": str(str_raw).strip() if str_raw else None,
            "plaats": str(pl_raw).strip() if pl_raw else None,
        }

    resultaten = []
    peildata = set()
    for row in rows[1:]:
        comp = extract_componenten(row)
        if not comp["adres_invoer"]:
            resultaten.append({
                "adres_invoer": "",
                "status": "Leeg",
                "wozWaarden": [],
                "adres_match": False,
                "adres_match_reden": "Rij leeg",
                "match_methode": "strict",
            })
            continue
        res = maak_resultaat_strict(
            adres_invoer=comp["adres_invoer"],
            postcode=comp["postcode"],
            huisnummer=comp["huisnummer"],
            huisletter=comp["huisletter"],
            toevoeging=comp["toevoeging"],
            straat_hint=comp["straat"],
        )
        for w in res.get("wozWaarden", []):
            if w.get("peildatum"):
                peildata.add(w["peildatum"])
        resultaten.append(res)
        time.sleep(0.15)  # voorkom rate-limit (60/min)

    peildata_sorted = sorted(peildata, reverse=True)

    wb_out = _bouw_excel(header, rows, resultaten, peildata_sorted)
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    naam = re.sub(r"\.xlsx?$", "", upload.filename, flags=re.I) + "_woz.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=naam,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --- Excel-stijlen ---
_RED_DARK = "1F4E79"
_RED_MID = "2E6BA8"
_GREY = "F1F4F8"
_BORDER = Side(style="thin", color="C8D2DD")
_BORDER_THICK = Side(style="medium", color=_RED_DARK)

_LABEL_FILL = {
    "A++++": "009639", "A+++": "009639", "A++": "009639", "A+": "009639",
    "A": "009639", "B": "4CAF50", "C": "8BC34A",
    "D": "FFC107", "E": "FF9800", "F": "FF5722", "G": "D32F2F",
}


def _bouw_excel(header_in, rows, resultaten, peildata_sorted) -> Workbook:
    """Bouwt een net opgemaakt resultaat-workbook met twee tabbladen."""
    wb = Workbook()

    # ---- Tab 1: Resultaten ----
    ws = wb.active
    ws.title = "Resultaten"

    oudste_peiljaar = peildata_sorted[-1][:4] if peildata_sorted else "oudste"

    invoer_kols = list(header_in) or ["adres"]
    locatie_kols = ["status", "match", "gevonden adres", "postcode", "woonplaats", "BAG-id"]
    bag_kols = ["bouwjaar", "opp. (m²)", "gebruiksdoel"]
    buurt_kols = ["buurt", "gem. WOZ buurt", "gem. inkomen", "% huur"]
    label_kols = ["energielabel", "gebouwtype", "geldig t/m"]
    mon_kols = ["rijksmonument"]
    afgeleid_kols = ["€/m²", "% 1jr", "% 5jr", f"% sinds {oudste_peiljaar}"]
    woz_kols = [p[:4] for p in peildata_sorted]

    sizes = {
        "in": len(invoer_kols), "loc": len(locatie_kols), "bag": len(bag_kols),
        "buurt": len(buurt_kols), "lab": len(label_kols), "mon": len(mon_kols),
        "afg": len(afgeleid_kols), "woz": len(woz_kols),
    }

    # Rij 1: groepheaders (merged)
    groepen = [
        ("Invoer", sizes["in"], "6B7785"),
        ("Adres & WOZ-object", sizes["loc"], _RED_DARK),
        ("BAG-gegevens", sizes["bag"], "8B5A2B"),
        ("Buurt (CBS 2023)", sizes["buurt"], "0F766E"),
        ("Energielabel (EP-online)", sizes["lab"], "1B7A3A"),
        ("Monument (RCE)", sizes["mon"], "7C3AED"),
        ("Trend", sizes["afg"], "B45309"),
        ("WOZ-waarden", sizes["woz"], _RED_MID),
    ]
    col = 1
    for naam, n, kleur in groepen:
        if n == 0:
            continue
        ws.cell(row=1, column=col, value=naam)
        if n > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n - 1)
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=kleur)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=_BORDER_THICK, left=_BORDER, right=_BORDER, bottom=_BORDER)
        col += n
    ws.row_dimensions[1].height = 24

    # Rij 2: kolomheaders
    headers_row2 = (invoer_kols + locatie_kols + bag_kols + buurt_kols
                    + label_kols + mon_kols + afgeleid_kols + woz_kols)
    for i, h in enumerate(headers_row2, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="1F2933", size=10)
        c.fill = PatternFill("solid", fgColor=_GREY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=_BORDER, right=_BORDER, bottom=_BORDER)
    ws.row_dimensions[2].height = 30

    # Data
    mismatched_rows = []
    for row, res in zip(rows[1:], resultaten):
        bag = res.get("bag") or {}
        ep = res.get("energielabel") or {}
        cbs = res.get("cbs") or {}
        mon = res.get("monument") or {}
        out_row = list(row)
        out_row += [None] * (sizes["in"] - len(out_row))
        out_row = out_row[:sizes["in"]]

        match_ok = res.get("adres_match")
        match_cell = "OK" if match_ok else (res.get("adres_match_reden") or "—")

        out_row += [
            res.get("status"),
            match_cell,
            res.get("weergavenaam"),
            res.get("postcode"),
            res.get("woonplaats"),
            res.get("adresseerbaarobject_id") or res.get("nummeraanduiding"),
        ]
        out_row += [bag.get("bouwjaar"), bag.get("gebruiksoppervlakte"), bag.get("gebruiksdoel")]
        out_row += [
            cbs.get("buurtnaam"),
            cbs.get("gem_woningwaarde"),
            cbs.get("gem_inkomen"),
            cbs.get("pct_huur"),
        ]
        out_row += [ep.get("energieklasse"), ep.get("gebouwtype"), ep.get("geldig_tot")]
        out_row += [mon.get("monumentnummer") or "—"]
        out_row += [
            res.get("woz_per_m2"),
            res.get("pct_1jr"),
            res.get("pct_5jr"),
            res.get("pct_sinds_oudst"),
        ]
        waarden_map = {w["peildatum"]: w["vastgesteldeWaarde"] for w in res.get("wozWaarden", [])}
        for p in peildata_sorted:
            out_row.append(waarden_map.get(p))
        ws.append(out_row)
        if match_ok is False:
            mismatched_rows.append(ws.max_row)

    max_row = ws.max_row
    max_col = len(headers_row2)

    # Borders + zebra
    for r in range(3, max_row + 1):
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = Border(left=_BORDER, right=_BORDER, bottom=_BORDER)
            if r % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="FAFBFC")

    # Mismatch-rijen oranje
    mismatch_fill = PatternFill("solid", fgColor="FFE8D6")
    for r in mismatched_rows:
        for c_idx in range(1, max_col + 1):
            ws.cell(row=r, column=c_idx).fill = mismatch_fill
        # Match-cel donkerder kleuren
        match_col = sizes["in"] + 2
        c = ws.cell(row=r, column=match_col)
        c.fill = PatternFill("solid", fgColor="FF9800")
        c.font = Font(bold=True, color="FFFFFF")

    # Energielabel-cel kleuren
    label_col = sizes["in"] + sizes["loc"] + sizes["bag"] + sizes["buurt"] + 1
    for r in range(3, max_row + 1):
        cell = ws.cell(row=r, column=label_col)
        lab = (cell.value or "").strip() if isinstance(cell.value, str) else ""
        if lab in _LABEL_FILL:
            cell.fill = PatternFill("solid", fgColor=_LABEL_FILL[lab])
            light = lab in ("D",)
            cell.font = Font(bold=True, color="1F2933" if light else "FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Monument-cel highlight als gevuld (niet —)
    mon_col = sizes["in"] + sizes["loc"] + sizes["bag"] + sizes["buurt"] + sizes["lab"] + 1
    for r in range(3, max_row + 1):
        cell = ws.cell(row=r, column=mon_col)
        if cell.value and cell.value != "—":
            cell.fill = PatternFill("solid", fgColor="7C3AED")
            cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # BAG-bouwjaar + opp formattering
    bouwjaar_col = sizes["in"] + sizes["loc"] + 1
    opp_col = sizes["in"] + sizes["loc"] + 2
    for r in range(3, max_row + 1):
        ws.cell(row=r, column=bouwjaar_col).number_format = "0"
        ws.cell(row=r, column=bouwjaar_col).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=opp_col).number_format = "0"
        ws.cell(row=r, column=opp_col).alignment = Alignment(horizontal="center")

    # CBS-kolommen formattering
    buurt_start = sizes["in"] + sizes["loc"] + sizes["bag"] + 1
    # buurtnaam | gem WOZ | gem inkomen | % huur
    for r in range(3, max_row + 1):
        ws.cell(row=r, column=buurt_start + 1).number_format = '"€" #,##0'
        ws.cell(row=r, column=buurt_start + 2).number_format = '"€" #,##0'
        ws.cell(row=r, column=buurt_start + 3).number_format = "0\"%\""
        ws.cell(row=r, column=buurt_start + 3).alignment = Alignment(horizontal="center")

    # Afgeleide kolommen: €/m², %1jr, %5jr, %sinds
    afg_start = sizes["in"] + sizes["loc"] + sizes["bag"] + sizes["buurt"] + sizes["lab"] + sizes["mon"] + 1
    for r in range(3, max_row + 1):
        ws.cell(row=r, column=afg_start).number_format = '"€" #,##0'
        for off in (1, 2, 3):
            c = ws.cell(row=r, column=afg_start + off)
            c.number_format = "0.0\"%\""
            c.alignment = Alignment(horizontal="right")
    # Conditional formatting op %-kolommen: groen positief, rood negatief
    for off in (1, 2, 3):
        col_letter = get_column_letter(afg_start + off)
        rng = f"{col_letter}3:{col_letter}{max_row}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=-20, start_color="E57373",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=50, end_color="81C784",
        ))

    # Euro-format + ColorScale voor WOZ-kolommen
    woz_start = sizes["in"] + sizes["loc"] + sizes["bag"] + sizes["buurt"] + sizes["lab"] + sizes["mon"] + sizes["afg"] + 1
    woz_end = woz_start + sizes["woz"] - 1
    for c_idx in range(woz_start, woz_end + 1):
        for r in range(3, max_row + 1):
            ws.cell(row=r, column=c_idx).number_format = '"€" #,##0'
        if max_row >= 3:
            rng = f"{get_column_letter(c_idx)}3:{get_column_letter(c_idx)}{max_row}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="DCE6F2",
                end_type="max", end_color="9BB8DC",
            ))

    # Kolombreedtes
    def set_width(col_letter, w):
        ws.column_dimensions[col_letter].width = w

    for i in range(1, sizes["in"] + 1):
        set_width(get_column_letter(i), 38)
    loc_widths = [11, 8, 38, 10, 16, 18]
    for i, w in enumerate(loc_widths):
        set_width(get_column_letter(sizes["in"] + 1 + i), w)
    bag_widths = [10, 10, 20]
    for i, w in enumerate(bag_widths):
        set_width(get_column_letter(sizes["in"] + sizes["loc"] + 1 + i), w)
    buurt_widths = [22, 14, 14, 9]
    for i, w in enumerate(buurt_widths):
        set_width(get_column_letter(sizes["in"] + sizes["loc"] + sizes["bag"] + 1 + i), w)
    lab_widths = [11, 22, 12]
    for i, w in enumerate(lab_widths):
        set_width(get_column_letter(sizes["in"] + sizes["loc"] + sizes["bag"] + sizes["buurt"] + 1 + i), w)
    set_width(get_column_letter(mon_col), 14)
    afg_widths = [10, 9, 9, 12]
    for i, w in enumerate(afg_widths):
        set_width(get_column_letter(afg_start + i), w)
    for i in range(sizes["woz"]):
        set_width(get_column_letter(woz_start + i), 12)

    ws.freeze_panes = ws.cell(row=3, column=sizes["in"] + 1)
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"

    # ---- Tab 2: Samenvatting ----
    ws2 = wb.create_sheet("Samenvatting")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 18

    title = ws2.cell(row=1, column=1, value="Samenvatting WOZ + energielabel")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=_RED_DARK)
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 28
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    totaal = len(resultaten)
    ok = sum(1 for r in resultaten if r.get("status") == "OK")
    niet_gevonden = sum(1 for r in resultaten if r.get("status") in ("Adres niet gevonden", "Geen nummeraanduiding"))
    geen_woz = totaal - ok - niet_gevonden

    woz_actueel = [r["wozWaarden"][0]["vastgesteldeWaarde"]
                   for r in resultaten if r.get("wozWaarden") and r["wozWaarden"][0].get("vastgesteldeWaarde")]
    gem_woz = int(sum(woz_actueel) / len(woz_actueel)) if woz_actueel else None
    min_woz = min(woz_actueel) if woz_actueel else None
    max_woz = max(woz_actueel) if woz_actueel else None

    labels_lijst = [(r.get("energielabel") or {}).get("energieklasse") for r in resultaten]
    labels_lijst = [l for l in labels_lijst if l]
    label_count = {}
    for l in labels_lijst:
        label_count[l] = label_count.get(l, 0) + 1

    # Bouwjaar uit BAG (autoritatief)
    bouwjaren = [(r.get("bag") or {}).get("bouwjaar") for r in resultaten]
    bouwjaren = [b for b in bouwjaren if isinstance(b, int) and b > 1500]
    gem_bj = int(sum(bouwjaren) / len(bouwjaren)) if bouwjaren else None
    oudste = min(bouwjaren) if bouwjaren else None
    nieuwste = max(bouwjaren) if bouwjaren else None

    opp = [(r.get("bag") or {}).get("gebruiksoppervlakte") for r in resultaten]
    opp = [o for o in opp if isinstance(o, (int, float)) and o > 0]
    gem_opp = round(sum(opp) / len(opp)) if opp else None
    min_opp = min(opp) if opp else None
    max_opp = max(opp) if opp else None

    laatste_peildatum = peildata_sorted[0] if peildata_sorted else "—"

    # Nieuw: WOZ per m²
    per_m2 = [r.get("woz_per_m2") for r in resultaten if r.get("woz_per_m2")]
    gem_per_m2 = int(sum(per_m2) / len(per_m2)) if per_m2 else None
    min_per_m2 = min(per_m2) if per_m2 else None
    max_per_m2 = max(per_m2) if per_m2 else None

    # Stijgingen
    p5 = [r.get("pct_5jr") for r in resultaten if r.get("pct_5jr") is not None]
    gem_p5 = round(sum(p5) / len(p5), 1) if p5 else None
    p_oudst = [r.get("pct_sinds_oudst") for r in resultaten if r.get("pct_sinds_oudst") is not None]
    gem_p_oudst = round(sum(p_oudst) / len(p_oudst), 1) if p_oudst else None

    # Mismatches
    mismatches = sum(1 for r in resultaten if r.get("adres_match") is False)

    # Monumenten
    monumenten = sum(1 for r in resultaten if (r.get("monument") or {}).get("monumentnummer"))

    blokken = [
        ("Aantal adressen", [
            ("Totaal verwerkt", totaal),
            ("Succesvol gevonden", ok),
            ("Adres-match-waarschuwingen", mismatches),
            ("Adres niet gevonden", niet_gevonden),
            ("Geen WOZ-waarde", geen_woz),
            ("Rijksmonumenten", monumenten),
        ]),
        (f"WOZ-waarde ({laatste_peildatum})", [
            ("Aantal met WOZ", len(woz_actueel)),
            ("Gemiddeld", gem_woz),
            ("Laagste", min_woz),
            ("Hoogste", max_woz),
        ]),
        ("WOZ per m² (BAG-opp.)", [
            ("Aantal met €/m²", len(per_m2)),
            ("Gemiddeld €/m²", gem_per_m2),
            ("Laagste €/m²", min_per_m2),
            ("Hoogste €/m²", max_per_m2),
        ]),
        ("WOZ-stijging (gemiddeld)", [
            ("% 5 jaar", gem_p5),
            (f"% sinds oudste peiljaar", gem_p_oudst),
        ]),
        ("Energielabels", [(f"Label {k}", v) for k, v in sorted(label_count.items())] or [("Geen labels gevonden", 0)]),
        ("Bouwjaar (BAG)", [
            ("Aantal met bouwjaar", len(bouwjaren)),
            ("Gemiddeld", gem_bj),
            ("Oudst", oudste),
            ("Nieuwst", nieuwste),
        ]),
        ("Gebruiksoppervlakte (BAG, m²)", [
            ("Aantal met oppervlakte", len(opp)),
            ("Gemiddeld", gem_opp),
            ("Kleinst", min_opp),
            ("Grootst", max_opp),
        ]),
    ]

    r = 3
    for titel, items in blokken:
        c = ws2.cell(row=r, column=1, value=titel)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=_RED_MID)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws2.row_dimensions[r].height = 20
        r += 1
        for k, v in items:
            ws2.cell(row=r, column=1, value=k).alignment = Alignment(indent=1)
            cv = ws2.cell(row=r, column=2, value=v)
            cv.alignment = Alignment(horizontal="right")
            if titel.startswith("WOZ-waarde") and k != "Aantal met WOZ":
                cv.number_format = '"€" #,##0'
            elif titel.startswith("WOZ per m²") and "€/m²" in k:
                cv.number_format = '"€" #,##0'
            elif titel.startswith("WOZ-stijging"):
                cv.number_format = "0.0\"%\""
            for col_idx in (1, 2):
                ws2.cell(row=r, column=col_idx).border = Border(bottom=_BORDER)
            r += 1
        r += 1  # lege tussen blokken

    return wb


@app.route("/api/template")
def api_template():
    """Genereer een lege voorbeeld-Excel met gestructureerde kolommen.

    Postcode en huisnummer zijn leidend (strikte matching). Straat en plaats
    zijn optioneel en alleen ter verificatie.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Adressen"
    kolommen = ["postcode", "huisnummer", "huisletter", "toevoeging", "straat", "plaats"]
    ws.append(kolommen)
    voorbeelden = [
        ["3012BK", 1, "D", "", "Witte de Withstraat", "Rotterdam"],
        ["1054GG", 70, "", "", "Vondelstraat", "Amsterdam"],
        ["3581WB", 1, "", "", "Bloemstraat", "Utrecht"],
    ]
    for v in voorbeelden:
        ws.append(v)

    # Header-styling
    breedtes = [11, 12, 10, 12, 26, 18]
    for col_idx, w in enumerate(breedtes, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Uitleg-tab
    ws2 = wb.create_sheet("Uitleg")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 90
    uitleg = [
        ("Kolom", "Toelichting"),
        ("postcode", "VERPLICHT — postcode in formaat 1234AB of '1234 AB'. Zonder postcode geen match."),
        ("huisnummer", "VERPLICHT — alleen het cijferdeel (bijv. 70)."),
        ("huisletter", "Optioneel — losse letter direct achter het huisnummer (bijv. 'D' bij '1D')."),
        ("toevoeging", "Optioneel — toevoeging na een streepje of spatie (bijv. '2' bij '70-2', 'hs' bij '70-hs')."),
        ("straat", "Optioneel — wordt alleen als hint gebruikt bij ambigue resultaten."),
        ("plaats", "Optioneel — informatief; matching gaat op postcode."),
        ("", ""),
        ("Hoe werkt het?", "Bij Excel-upload gebruikt de tool postcode + huisnummer als sleutel. "
                          "Bij geen exacte match krijgt die rij in de output 'geen match' i.p.v. een gokje."),
        ("Vrije adres-kolom?", "Een kolom met naam 'adres' (vrije tekst) mag ook. De tool parst dan zelf "
                                "postcode + huisnummer uit die tekst. Werkt alleen betrouwbaar als beide erin staan."),
    ]
    for r_idx, (k, v) in enumerate(uitleg, start=1):
        ws2.cell(row=r_idx, column=1, value=k)
        ws2.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            for col in (1, 2):
                c = ws2.cell(row=r_idx, column=col)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E79")
                c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            ws2.cell(row=r_idx, column=1).font = Font(bold=True)
            for col in (1, 2):
                ws2.cell(row=r_idx, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top",
                )
        if r_idx > 1:
            ws2.row_dimensions[r_idx].height = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="woz_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Realworks-koppeling
# ---------------------------------------------------------------------------


@app.route("/realworks")
def realworks_page():
    """HTML-pagina met overzicht van Realworks-objecten."""
    return render_template(
        "realworks.html",
        geconfigureerd=realworks_service.is_geconfigureerd(),
    )


@app.route("/api/realworks")
def api_realworks():
    """JSON-endpoint: één pagina Realworks-objecten."""
    try:
        vanaf = int(request.args.get("vanaf", 0))
        aantal = int(request.args.get("aantal", 100))
    except ValueError:
        return jsonify({"error": "Parameter 'vanaf' en 'aantal' moeten getallen zijn."}), 400
    status = (request.args.get("status") or "").strip() or None

    try:
        page = realworks_service.haal_objecten(vanaf=vanaf, aantal=aantal, status=status)
    except realworks_service.RealworksError as e:
        return jsonify({"error": str(e)}), e.status_code or 502
    return jsonify(page)


@app.route("/api/realworks/alle")
def api_realworks_alle():
    """JSON-endpoint: alle objecten via auto-paginatie (gecapt)."""
    status = (request.args.get("status") or "").strip() or None
    try:
        max_obj = int(request.args.get("max", 500))
    except ValueError:
        max_obj = 500
    try:
        result = realworks_service.haal_alle_objecten(status=status, max_objecten=max_obj)
    except realworks_service.RealworksError as e:
        return jsonify({"error": str(e)}), e.status_code or 502
    return jsonify(result)


@app.route("/api/realworks/verrijk", methods=["POST"])
def api_realworks_verrijk():
    """Neemt een lijst Realworks-objecten en draait WOZ/BAG/label/CBS/monument.

    Gebruikt strikte adresmatching op postcode + huisnummer (+ optioneel
    toevoeging) zodat dezelfde straatnaam in een andere stad nooit per ongeluk
    gekoppeld wordt.
    """
    payload = request.get_json(silent=True) or {}
    objecten = payload.get("objecten") or []
    if not objecten:
        return jsonify({"error": "Geen objecten meegegeven."}), 400

    verrijkt = []
    for obj in objecten:
        adres_invoer = realworks_service.adres_string(obj) or ""
        postcode = obj.get("postcode")
        huisnr_raw = obj.get("huisnummer")
        toev_raw = obj.get("huisnummerToevoeging")
        straat_hint = obj.get("straat")

        # Realworks levert huisnummer soms als int, soms als string ("70a")
        if huisnr_raw is None or str(huisnr_raw).strip() == "":
            huisnummer, huisletter, toevoeging = None, None, None
        else:
            huisnummer, huisletter, toevoeging = _split_huisnummer_componenten(
                str(huisnr_raw)
            )
        if toev_raw and not toevoeging:
            toevoeging = str(toev_raw).strip() or None

        if not adres_invoer:
            res = {
                "adres_invoer": "",
                "status": "Leeg",
                "wozWaarden": [],
                "adres_match": False,
                "match_methode": "strict",
            }
        else:
            res = maak_resultaat_strict(
                adres_invoer=adres_invoer,
                postcode=postcode,
                huisnummer=huisnummer,
                huisletter=huisletter,
                toevoeging=toevoeging,
                straat_hint=straat_hint,
            )
        verrijkt.append({"realworks": obj, "resultaat": res})
        time.sleep(0.15)
    return jsonify({"verrijkt": verrijkt, "aantal": len(verrijkt)})


if __name__ == "__main__":
    print("\n  WOZ-tool gestart op  http://127.0.0.1:5005\n")
    app.run(host="127.0.0.1", port=5005, debug=False)
